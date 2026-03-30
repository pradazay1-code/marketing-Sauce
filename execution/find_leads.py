#!/usr/bin/env python3
"""
Lead Finder — Find small businesses without websites in MA, RI, CT.

Sources:
  1. YP.com (Yellow Pages) — category + city search
  2. Yelp — category + location search

Output: Excel spreadsheet with business details
"""

import os
import re
import sys
import csv
import time
import random
import logging
from datetime import datetime
from urllib.parse import quote_plus, urljoin

import requests
from bs4 import BeautifulSoup

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# --- Config ---
HEADERS = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# Delay between requests (seconds) — be respectful
MIN_DELAY = 2
MAX_DELAY = 4

# Known chains to skip
CHAINS = {
    "mcdonald", "burger king", "wendy", "subway", "dunkin", "starbucks",
    "taco bell", "pizza hut", "domino", "papa john", "chick-fil-a",
    "popeyes", "kfc", "arby", "sonic", "chipotle", "panera",
    "applebee", "olive garden", "red lobster", "ihop", "denny",
    "walmart", "target", "walgreens", "cvs", "rite aid",
    "home depot", "lowe", "dollar tree", "dollar general",
    "autozone", "o'reilly", "advance auto", "jiffy lube",
    "supercuts", "great clips", "sport clips",
    "anytime fitness", "planet fitness", "orangetheory",
}

# Business categories to search
CATEGORIES = [
    "restaurants", "hair salons", "barber shops", "auto repair",
    "landscaping", "cleaning services", "contractors", "plumbers",
    "electricians", "painters", "photography studios", "catering",
    "bakeries", "food trucks", "tattoo shops", "pet grooming",
    "tutoring services", "handyman services", "florists",
    "fitness studios", "yoga studios", "daycare", "event planning",
    "nail salons", "massage therapy", "moving companies",
    "pressure washing", "roofing", "fencing contractors",
]

# Cities to search — mix of sizes across MA, RI, CT
CITIES = {
    "MA": [
        "New Bedford", "Fall River", "Brockton", "Taunton", "Attleboro",
        "Worcester", "Springfield", "Lowell", "Lawrence", "Haverhill",
        "Fitchburg", "Leominster", "Pittsfield", "Holyoke", "Chicopee",
        "Revere", "Lynn", "Salem", "Gloucester", "Marlborough",
    ],
    "RI": [
        "Providence", "Warwick", "Cranston", "Pawtucket", "East Providence",
        "Woonsocket", "Newport", "Central Falls", "West Warwick",
        "Cumberland", "North Providence", "Bristol", "Westerly",
    ],
    "CT": [
        "Hartford", "New Haven", "Bridgeport", "Waterbury", "Stamford",
        "Norwalk", "Danbury", "New Britain", "Meriden", "Middletown",
        "Norwich", "Torrington", "Shelton", "Bristol", "Windham",
    ],
}

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)


def polite_delay():
    """Sleep a random interval to avoid hammering servers."""
    time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))


def is_chain(name):
    """Check if a business name looks like a national chain."""
    lower = name.lower()
    return any(chain in lower for chain in CHAINS)


def clean_phone(raw):
    """Extract a 10-digit US phone number from text."""
    if not raw:
        return ""
    digits = re.sub(r'\D', '', raw)
    if len(digits) == 11 and digits.startswith('1'):
        digits = digits[1:]
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    return raw.strip()


# ---------------------------------------------------------------------------
#  Source 1: YP.com (Yellow Pages)
# ---------------------------------------------------------------------------

def scrape_yp(category, city, state):
    """Scrape YP.com for businesses in a category + city."""
    leads = []
    search_term = quote_plus(category)
    location = quote_plus(f"{city}, {state}")
    url = f"https://www.yellowpages.com/search?search_terms={search_term}&geo_location_terms={location}"

    try:
        log.info(f"  YP: {category} in {city}, {state}")
        resp = requests.get(url, headers=HEADERS, timeout=15)
        if resp.status_code != 200:
            log.warning(f"  YP returned {resp.status_code}")
            return leads

        soup = BeautifulSoup(resp.text, "lxml")
        results = soup.select("div.result")
        if not results:
            results = soup.select("div.v-card")
        if not results:
            results = soup.select("div.search-results div.srp-listing")

        for r in results:
            try:
                # Business name
                name_el = r.select_one("a.business-name") or r.select_one("h2 a") or r.select_one(".n a")
                if not name_el:
                    continue
                name = name_el.get_text(strip=True)
                if not name or is_chain(name):
                    continue

                # Phone
                phone_el = r.select_one("div.phones") or r.select_one(".phone")
                phone = clean_phone(phone_el.get_text(strip=True)) if phone_el else ""

                # Address
                addr_el = r.select_one("div.adr") or r.select_one(".adr")
                street = ""
                locality = city
                if addr_el:
                    street_el = addr_el.select_one("span.street-address")
                    locality_el = addr_el.select_one("span.locality")
                    street = street_el.get_text(strip=True) if street_el else ""
                    locality = locality_el.get_text(strip=True) if locality_el else city

                # Website check
                website_el = r.select_one("a.track-visit-website") or r.select_one("a[href*='website']")
                has_website = bool(website_el)
                website_url = ""
                if website_el:
                    website_url = website_el.get("href", "")

                # Description / categories
                cats_el = r.select_one("div.categories") or r.select_one(".cats")
                desc = cats_el.get_text(strip=True) if cats_el else category

                if not has_website:
                    leads.append({
                        "Business Name": name,
                        "Category": category.title(),
                        "Description": desc,
                        "Phone": phone,
                        "Owner Name": "",
                        "Address": street,
                        "City": locality,
                        "State": state,
                        "Has Website": "No",
                        "Website URL": "",
                        "Source": "YellowPages",
                        "Date Found": datetime.now().strftime("%Y-%m-%d"),
                        "Notes": "",
                    })
            except Exception as e:
                log.debug(f"  Error parsing YP result: {e}")
                continue

    except requests.RequestException as e:
        log.warning(f"  YP request failed: {e}")

    return leads


# ---------------------------------------------------------------------------
#  Source 2: Yelp
# ---------------------------------------------------------------------------

def scrape_yelp(category, city, state):
    """Scrape Yelp for businesses in a category + city."""
    leads = []
    search_term = quote_plus(category)
    location = quote_plus(f"{city}, {state}")
    url = f"https://www.yelp.com/search?find_desc={search_term}&find_loc={location}"

    try:
        log.info(f"  Yelp: {category} in {city}, {state}")
        resp = requests.get(url, headers=HEADERS, timeout=15)
        if resp.status_code != 200:
            log.warning(f"  Yelp returned {resp.status_code}")
            return leads

        soup = BeautifulSoup(resp.text, "lxml")

        # Yelp uses various div structures — try multiple selectors
        results = soup.select("[data-testid='serp-ia-card']")
        if not results:
            results = soup.select("div.container__09f24__FeTO6")
        if not results:
            # Fallback: look for business name links
            biz_links = soup.select("a[href*='/biz/']")
            for link in biz_links:
                name = link.get_text(strip=True)
                if not name or len(name) < 3 or is_chain(name):
                    continue
                # Check if there's a website link nearby
                parent = link.find_parent("div")
                if parent:
                    website_link = parent.select_one("a[href*='biz_redir']")
                    if not website_link:
                        phone_el = parent.select_one("[class*='phone']")
                        phone = clean_phone(phone_el.get_text(strip=True)) if phone_el else ""

                        leads.append({
                            "Business Name": name,
                            "Category": category.title(),
                            "Description": f"{category.title()} in {city}, {state}",
                            "Phone": phone,
                            "Owner Name": "",
                            "Address": "",
                            "City": city,
                            "State": state,
                            "Has Website": "No",
                            "Website URL": "",
                            "Source": "Yelp",
                            "Date Found": datetime.now().strftime("%Y-%m-%d"),
                            "Notes": "",
                        })
            return leads

        for r in results:
            try:
                name_el = r.select_one("a[href*='/biz/']")
                if not name_el:
                    continue
                name = name_el.get_text(strip=True)
                if not name or is_chain(name):
                    continue

                # Check for website
                website_el = r.select_one("a[href*='biz_redir']")
                has_website = bool(website_el)

                phone_el = r.select_one("[class*='phone']")
                phone = clean_phone(phone_el.get_text(strip=True)) if phone_el else ""

                if not has_website:
                    leads.append({
                        "Business Name": name,
                        "Category": category.title(),
                        "Description": f"{category.title()} in {city}, {state}",
                        "Phone": phone,
                        "Owner Name": "",
                        "Address": "",
                        "City": city,
                        "State": state,
                        "Has Website": "No",
                        "Website URL": "",
                        "Source": "Yelp",
                        "Date Found": datetime.now().strftime("%Y-%m-%d"),
                        "Notes": "",
                    })
            except Exception as e:
                log.debug(f"  Error parsing Yelp result: {e}")
                continue

    except requests.RequestException as e:
        log.warning(f"  Yelp request failed: {e}")

    return leads


# ---------------------------------------------------------------------------
#  Source 3: MA Secretary of State — New Business Filings
# ---------------------------------------------------------------------------

def scrape_ma_sos():
    """
    Scrape recent MA business filings from the Secretary of State website.
    These are newly registered LLCs/Corps — likely startups without websites.
    """
    leads = []
    log.info("Checking MA Secretary of State for new business filings...")

    # The MA SOS site uses a search form — we'll search for recent entities
    url = "https://corp.sec.state.ma.us/corpweb/CorpSearch/CorpSearch.aspx"

    try:
        session = requests.Session()
        session.headers.update(HEADERS)

        # Get the page first to extract form tokens
        resp = session.get(url, timeout=15)
        if resp.status_code != 200:
            log.warning(f"  MA SOS returned {resp.status_code}")
            return leads

        soup = BeautifulSoup(resp.text, "lxml")

        # Extract ASP.NET form fields
        viewstate = soup.select_one("#__VIEWSTATE")
        viewstate_gen = soup.select_one("#__VIEWSTATEGENERATOR")
        event_validation = soup.select_one("#__EVENTVALIDATION")

        if not viewstate:
            log.warning("  Could not find form tokens on MA SOS page")
            return leads

        # Search for recently filed entities (LLC is most common for startups)
        form_data = {
            "__VIEWSTATE": viewstate.get("value", ""),
            "__VIEWSTATEGENERATOR": viewstate_gen.get("value", "") if viewstate_gen else "",
            "__EVENTVALIDATION": event_validation.get("value", "") if event_validation else "",
            "ctl00$MainContent$txtEntityName": "",
            "ctl00$MainContent$ddlEntityType": "LLC",
            "ctl00$MainContent$btnSearch": "Search",
        }

        polite_delay()
        resp = session.post(url, data=form_data, timeout=15)

        if resp.status_code == 200:
            soup = BeautifulSoup(resp.text, "lxml")
            rows = soup.select("table#MainContent_SearchResults tr")[1:]  # Skip header

            for row in rows[:30]:  # Limit to first 30 results
                cells = row.select("td")
                if len(cells) >= 3:
                    name = cells[0].get_text(strip=True)
                    entity_type = cells[1].get_text(strip=True) if len(cells) > 1 else "LLC"
                    date_filed = cells[2].get_text(strip=True) if len(cells) > 2 else ""

                    if name and not is_chain(name):
                        # Clean up LLC/Inc from display
                        display_name = name.replace(" LLC", "").replace(" Inc", "").replace(" Corp", "").strip()

                        leads.append({
                            "Business Name": display_name,
                            "Category": "New Business Filing",
                            "Description": f"{entity_type} — Filed: {date_filed}",
                            "Phone": "",
                            "Owner Name": "",
                            "Address": "",
                            "City": "",
                            "State": "MA",
                            "Has Website": "Unknown",
                            "Website URL": "",
                            "Source": "MA Secretary of State",
                            "Date Found": datetime.now().strftime("%Y-%m-%d"),
                            "Notes": f"Entity type: {entity_type}. Filed: {date_filed}. Likely no website yet.",
                        })

    except requests.RequestException as e:
        log.warning(f"  MA SOS request failed: {e}")

    return leads


# ---------------------------------------------------------------------------
#  Deduplication
# ---------------------------------------------------------------------------

def deduplicate(leads):
    """Remove duplicate businesses by name + city."""
    seen = set()
    unique = []
    for lead in leads:
        key = (lead["Business Name"].lower().strip(), lead["City"].lower().strip())
        if key not in seen:
            seen.add(key)
            unique.append(lead)
    return unique


# ---------------------------------------------------------------------------
#  Output
# ---------------------------------------------------------------------------

COLUMNS = [
    "Business Name", "Category", "Description", "Phone", "Owner Name",
    "Address", "City", "State", "Has Website", "Website URL",
    "Source", "Date Found", "Notes",
]


def save_to_excel(leads, filepath):
    """Save leads to a formatted Excel spreadsheet."""
    if not HAS_OPENPYXL:
        log.warning("openpyxl not installed — falling back to CSV")
        save_to_csv(leads, filepath.replace(".xlsx", ".csv"))
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Header styling
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data
    row_fill_even = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")
    data_font = Font(name="Arial", size=10)

    for row_idx, lead in enumerate(leads, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=lead.get(col_name, ""))
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = row_fill_even

    # Column widths
    col_widths = {
        "A": 30, "B": 18, "C": 35, "D": 16, "E": 20,
        "F": 30, "G": 16, "H": 8, "I": 12, "J": 30,
        "K": 18, "L": 12, "M": 30,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:M{len(leads) + 1}"

    wb.save(filepath)
    log.info(f"Saved {len(leads)} leads to {filepath}")


def save_to_csv(leads, filepath):
    """Fallback: save as CSV."""
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(leads)
    log.info(f"Saved {len(leads)} leads to {filepath}")


# ---------------------------------------------------------------------------
#  Main
# ---------------------------------------------------------------------------

def main():
    all_leads = []
    target_count = 50

    log.info("=" * 60)
    log.info("  LEAD FINDER — Small Businesses Without Websites")
    log.info(f"  Target: {target_count}+ leads in MA, RI, CT")
    log.info("=" * 60)

    # --- Source 1: Yellow Pages ---
    log.info("\n--- Source: Yellow Pages ---")
    # Pick a random subset of categories and cities to avoid too many requests
    categories_sample = random.sample(CATEGORIES, min(15, len(CATEGORIES)))

    for state, cities in CITIES.items():
        cities_sample = random.sample(cities, min(5, len(cities)))
        for city in cities_sample:
            for category in categories_sample:
                if len(all_leads) >= target_count * 2:
                    break
                new_leads = scrape_yp(category, city, state)
                all_leads.extend(new_leads)
                if new_leads:
                    log.info(f"    Found {len(new_leads)} leads")
                polite_delay()
            if len(all_leads) >= target_count * 2:
                break
        if len(all_leads) >= target_count * 2:
            break

    # --- Source 2: Yelp ---
    log.info("\n--- Source: Yelp ---")
    for state, cities in CITIES.items():
        cities_sample = random.sample(cities, min(3, len(cities)))
        for city in cities_sample:
            cats_sample = random.sample(CATEGORIES, min(8, len(CATEGORIES)))
            for category in cats_sample:
                if len(all_leads) >= target_count * 3:
                    break
                new_leads = scrape_yelp(category, city, state)
                all_leads.extend(new_leads)
                if new_leads:
                    log.info(f"    Found {len(new_leads)} leads")
                polite_delay()
            if len(all_leads) >= target_count * 3:
                break
        if len(all_leads) >= target_count * 3:
            break

    # --- Source 3: MA Secretary of State ---
    log.info("\n--- Source: MA Secretary of State ---")
    sos_leads = scrape_ma_sos()
    all_leads.extend(sos_leads)
    log.info(f"  Found {len(sos_leads)} new business filings")

    # --- Deduplicate ---
    log.info(f"\nTotal raw leads: {len(all_leads)}")
    all_leads = deduplicate(all_leads)
    log.info(f"After deduplication: {len(all_leads)}")

    if not all_leads:
        log.error("No leads found! Check network connection and try again.")
        sys.exit(1)

    # --- Save ---
    output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "clients", "leads")
    os.makedirs(output_dir, exist_ok=True)

    date_str = datetime.now().strftime("%Y-%m-%d")
    xlsx_path = os.path.join(output_dir, f"leads_{date_str}.xlsx")
    csv_path = os.path.join(output_dir, f"leads_{date_str}.csv")

    save_to_excel(all_leads, xlsx_path)
    save_to_csv(all_leads, csv_path)

    # --- Summary ---
    log.info("\n" + "=" * 60)
    log.info("  RESULTS SUMMARY")
    log.info("=" * 60)
    log.info(f"  Total unique leads: {len(all_leads)}")

    by_state = {}
    by_source = {}
    for lead in all_leads:
        st = lead["State"]
        src = lead["Source"]
        by_state[st] = by_state.get(st, 0) + 1
        by_source[src] = by_source.get(src, 0) + 1

    log.info("  By state:")
    for st, count in sorted(by_state.items()):
        log.info(f"    {st}: {count}")
    log.info("  By source:")
    for src, count in sorted(by_source.items()):
        log.info(f"    {src}: {count}")
    log.info(f"\n  Spreadsheet: {xlsx_path}")
    log.info(f"  CSV backup:  {csv_path}")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
