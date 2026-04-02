#!/usr/bin/env python3
"""Create a leads spreadsheet with 15 MA businesses that need websites."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "MA Business Leads"

# Styling
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
alt_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

headers = [
    "No.", "Business Name", "Type", "City", "State",
    "Address", "Phone", "Online Presence",
    "Website Status", "Notes", "Priority"
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = border

# 15 verified leads from research
leads = [
    [1, "Torres Family Barber Shop", "Barbershop", "Westborough", "MA",
     "72 East Main St, 2nd Floor", "N/A - New 2025",
     "None - just opened May 2025", "No Website",
     "Brand new barbershop opened May 2025. Owner Jose Torres has 8 yrs experience in the area. Located above Swanski's Bakery.", "HIGH"],

    [2, "Thela Indian Street Food", "Restaurant", "Westborough", "MA",
     "72 East Main St, Lower Level", "N/A - New 2025",
     "None found", "No Website",
     "New Indian street food shop in Westborough, opened 2025. Located below Swanski's Bakery on East Main St.", "HIGH"],

    [3, "Rancho Alegre", "Mexican Restaurant", "Westborough", "MA",
     "Former Kenzo's Steakhouse, East Main St", "N/A - Opening 2025",
     "None found", "No Website",
     "New Mexican restaurant converting former Kenzo's Steakhouse space. In process of opening 2025.", "HIGH"],

    [4, "Ace of Babes Vintage", "Vintage Clothing Shop", "Worcester", "MA",
     "Portland Street, Downtown", "N/A - New 2025",
     "Possibly social media only", "No Website",
     "Independently owned vintage clothing shop. Newly opened in downtown Worcester on Portland St.", "HIGH"],

    [5, "Boom Jewels", "Jewelry Boutique", "Worcester", "MA",
     "Main Street, Owl Shop Building", "N/A - New 2025",
     "Unknown", "No Website",
     "New jewelry boutique inside the historic Owl Shop Building on Main St downtown Worcester.", "HIGH"],

    [6, "Monalessa's Kitchen", "Indian Restaurant & Grill", "Fall River", "MA",
     "756 Brayton Ave (former Ed's Famous BBQ)", "N/A - Opening 2025",
     "None found", "No Website",
     "New Indian restaurant/grill opening in Fall River South End. Owner Md Shahidul Islam also owns NY Fresh Bagel.", "HIGH"],

    [7, "The Counting House", "Cocktail Bar & Restaurant", "Fall River", "MA",
     "Near Durfee Mills, Downtown", "N/A - Opened Jan 2025",
     "Possibly social media only", "No Website",
     "Craft cocktails, espresso martinis, small plates with Italian & Portuguese influences. Opened Jan 2, 2025.", "MEDIUM"],

    [8, "Morgan's Cantina", "Mexican Restaurant", "Fall River", "MA",
     "1 Ferry St, Waterfront", "N/A - Opening 2025",
     "None found", "No Website",
     "Handmade Mexican cuisine & craft cocktails. Waterfront views of Mount Hope Bay. Owners Michael & Nicole Lund.", "HIGH"],

    [9, "Slades Ferry Grille", "Restaurant", "Somerset", "MA",
     "Former Magoni's Ferry Landing", "N/A - Opening 2025",
     "None found", "No Website",
     "Reopening iconic 70+ year old restaurant space as Slades Ferry Grille after extensive renovations.", "MEDIUM"],

    [10, "Surfside Seafood", "Seafood Restaurant", "Orleans", "MA",
     "18 Old Colony Way", "N/A - New 2025",
     "None found", "No Website",
     "Run by two fishermen. Fresh Cape Cod seafood. Newly opened in Orleans on Cape Cod.", "HIGH"],

    [11, "Liberty Artisanal Bakery", "Bakery & Coffee", "Chatham", "MA",
     "1223 Main St", "N/A - New 2025",
     "Unknown", "No Website",
     "Artisan breads, pastries, coffee & espresso bar. Newly opened on Main St in Chatham.", "HIGH"],

    [12, "Cape Life Brand Company", "Apparel & Accessories", "Harwich Port", "MA",
     "537 Main St", "N/A - New 2025",
     "Unknown", "No Website",
     "Cape-themed apparel and accessories with stylish designs. New retail shop in Harwich Port.", "MEDIUM"],

    [13, "Mac Daddies", "Comfort Food Restaurant", "Haverhill", "MA",
     "72 South Main St", "N/A - Relocated 2025",
     "Possibly Facebook only", "No Website",
     "Creative mac-n-cheese comfort food. Relocated to bigger space. Opening 2nd location in Amesbury Jan 2026.", "MEDIUM"],

    [14, "WCP Slice House", "Pizza Restaurant", "Lowell", "MA",
     "Central Street, Downtown", "N/A - Opened Oct 2025",
     "Unknown", "No Website",
     "South Shore bar pizza, thin crust, Sicilian, New Haven style, GF options. Opened Oct 2025 in downtown Lowell.", "MEDIUM"],

    [15, "A New Chapter", "Independent Bookstore", "West Springfield", "MA",
     "Downtown West Springfield", "N/A - Opened Oct 2025",
     "None found", "No Website",
     "Independent bookstore opened Oct 7, 2025 by Scott Szaban. Grand opening Nov 7-9, 2025.", "HIGH"],
]

for i, lead in enumerate(leads):
    row = i + 2
    for col, val in enumerate(lead, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if i % 2 == 1:
            cell.fill = alt_fill

# Column widths
widths = [5, 28, 22, 18, 6, 35, 18, 22, 14, 55, 10]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# Freeze top row
ws.freeze_panes = "A2"

# Save
output = f"/home/user/marketing-Sauce/clients/leads/leads_{date.today().strftime('%Y-%m-%d')}.xlsx"
wb.save(output)
print(f"Saved: {output}")
print(f"Total leads: {len(leads)}")
